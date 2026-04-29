# -*- coding: utf-8 -*-
"""
Excel Export Module for Tapu Okuyucu.
Exports şerh entries to Excel format with optional column filtering.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# All available columns with their display names and entry keys
ALL_COLUMNS = [
    {"key": "type", "label": "Tür"},
    {"key": "icra_dairesi", "label": "İcra Dairesi"},
    {"key": "dosya_no", "label": "Dosya No"},
    {"key": "tarih", "label": "Tarih"},
    {"key": "bedel", "label": "Bedel (TL)"},
    {"key": "alacakli", "label": "Alacaklı"},
    {"key": "yevmiye_tarih", "label": "Yevmiye Tarih"},
    {"key": "yevmiye_no", "label": "Yevmiye No"},
]

def _filter_columns(selected_columns: list = None) -> list:
    """Filter columns based on user selection. Returns list of {key, label} dicts."""
    if not selected_columns:
        return ALL_COLUMNS
    return [col for col in ALL_COLUMNS if col["key"] in selected_columns]


def export_to_excel(grouped_entries: dict, malik_name: str, tapu_date: str,
                    output_path: str, selected_columns: list = None) -> str:
    """
    Export grouped şerh entries to an Excel file.
    grouped_entries: dict with İcra Dairesi as key and list of entries as value
    selected_columns: optional list of column keys to include (e.g. ["type", "dosya_no"])
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Şerh Listesi"

    # Filter columns
    columns = _filter_columns(selected_columns)
    # For grouped export, exclude icra_dairesi from columns (it's the group header)
    group_columns = [col for col in columns if col["key"] != "icra_dairesi"]
    num_cols = max(len(group_columns), 1)

    # Styles
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    col_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"Tapu Şerh Listesi - {malik_name}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # Date info
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"Tapu Kayıt Tarihi: {tapu_date}")
    cell.font = Font(name="Calibri", size=11, italic=True)
    cell.alignment = Alignment(horizontal="center")
    row += 2

    for dairesi, entries in grouped_entries.items():
        # İcra Dairesi header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=f"📁 {dairesi}")
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="left")
        row += 1

        # Column headers
        for col_idx, col_def in enumerate(group_columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_def["label"])
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # Data rows
        for entry in entries:
            for col_idx, col_def in enumerate(group_columns, 1):
                val = entry.get(col_def["key"], "")
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            row += 1

        row += 1  # Empty row between groups

    # Auto-adjust column widths
    for col_idx in range(1, num_cols + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 15
        for r in range(1, row):
            cell_value = ws.cell(row=r, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)
    return output_path


def export_comparison_to_excel(comparison: dict, malik_name: str,
                               old_date: str, new_date: str,
                               output_path: str, selected_columns: list = None) -> str:
    """Export comparison results to a single unified Excel sheet, grouped by İcra Dairesi."""
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = "Karşılaştırma Raporu"

    # Merge all entries and attach status
    all_entries = []
    for entry in comparison.get("removed", []):
        e = dict(entry)
        e["_status"] = "removed"
        e["_durum_text"] = "Kaldırıldı"
        all_entries.append(e)

    for entry in comparison.get("remaining", []):
        e = dict(entry)
        e["_status"] = "remaining"
        e["_durum_text"] = "Devam Ediyor"
        all_entries.append(e)

    for entry in comparison.get("added", []):
        e = dict(entry)
        e["_status"] = "added"
        e["_durum_text"] = "Yeni Eklendi"
        all_entries.append(e)

    # Group by İcra Dairesi
    grouped = defaultdict(list)
    for entry in all_entries:
        dairesi = entry.get("icra_dairesi", "Bilinmeyen Daire")
        if not dairesi.strip():
            dairesi = "Bilinmeyen Daire"
        grouped[dairesi].append(entry)

    # Setup Columns
    columns = _filter_columns(selected_columns)
    # Exclude icra_dairesi from columns since it's the group header
    group_columns = [col for col in columns if col["key"] != "icra_dairesi"]
    # Add Durum column at the end
    group_columns.append({"key": "_durum_text", "label": "Durum"})

    num_cols = max(len(group_columns), 1)

    # Styles
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    col_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="Calibri", size=10)

    # Status Colors
    fill_removed = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    fill_added = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # Light Green
    fill_remaining = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # White

    font_removed = Font(name="Calibri", size=10, color="990000", strike=True) # Dark red, strikethrough

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"Tapu Şerh Karşılaştırma Raporu - {malik_name}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # Date info
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"Eski Tarih: {old_date}  |  Yeni Tarih: {new_date}")
    cell.font = Font(name="Calibri", size=11, italic=True)
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # Summary
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"ÖZET: {len(comparison.get('removed', []))} Kaldırılan, {len(comparison.get('remaining', []))} Devam Eden, {len(comparison.get('added', []))} Yeni Eklenen")
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # Write Groups
    for dairesi in sorted(grouped.keys()):
        entries = grouped[dairesi]

        # İcra Dairesi header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=f"📁 {dairesi}")
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="left")
        row += 1

        # Column headers
        for col_idx, col_def in enumerate(group_columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_def["label"])
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # Data rows
        for entry in entries:
            status = entry.get("_status")
            current_fill = fill_remaining
            current_font = data_font

            if status == "removed":
                current_fill = fill_removed
                current_font = font_removed
            elif status == "added":
                current_fill = fill_added

            for col_idx, col_def in enumerate(group_columns, 1):
                val = entry.get(col_def["key"], "")
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = current_font
                cell.fill = current_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            row += 1

        row += 1  # Empty row between groups

    # Auto-adjust column widths
    for col_idx in range(1, num_cols + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 15
        for r in range(1, row):
            cell_value = ws.cell(row=r, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)
    return output_path
